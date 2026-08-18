#!/usr/bin/env python3
"""
試験項目書とエビデンスの整合性チェックシステム

このスクリプトは以下の処理を実行します:
1. 試験項目書のExcelファイルを読み込む
2. エビデンスのExcelファイル(ST-002シート)を読み込む
3. 試験観点とエビデンスを比較し、OK/NG判定を実施
4. 画像ファイルが含まれている場合は「※目視で確認してください」と記載
5. 実施日と実施者を自動入力
6. 結果を試験項目書に記載して保存
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import re

# ファイルパス設定
TEST_ITEMS_FILE = "試験項目書_ST_Ne-OpS名整合性チェック.xlsx"
EVIDENCE_FILE = "ST工ビデンス_Ne-OpS名整合性チェック.xlsx"

# シート名
TEST_ITEMS_SHEET = "試験項目書"
EVIDENCE_SHEET = "ST-002"

# 実施者名
IMPLEMENTER_NAME = "湊"

# 画像ファイルの拡張子
IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff'}

def extract_image_references(text):
    """
    テキストから画像ファイルの参照を抽出
    
    Args:
        text: 検査対象のテキスト
    
    Returns:
        画像ファイルが見つかったかどうかのブール値
    """
    if not text:
        return False
    
    # ファイルパスのパターンを検索
    file_pattern = r'([^\s/\\]+\.(png|jpg|jpeg|gif|bmp|tiff))'
    matches = re.findall(file_pattern, str(text), re.IGNORECASE)
    
    return len(matches) > 0


def check_evidence_coverage(expected_content, evidence_data):
    """
    期待値の確認内容とエビデンスをチェック
    
    Args:
        expected_content: 期待値の確認内容
        evidence_data: エビデンスデータ（行のリスト）
    
    Returns:
        (result, has_image_ref): (OK/NG, 画像参照の有無)
    """
    if not expected_content:
        return "NG", False
    
    # エビデンスから該当する行を検索
    expected_str = str(expected_content).lower().strip()
    has_image = False
    
    for evidence_row in evidence_data:
        # 行のすべてのセルをチェック
        for cell_value in evidence_row:
            if not cell_value:
                continue
            
            cell_str = str(cell_value).lower()
            
            # 画像ファイル参照をチェック
            if extract_image_references(cell_str):
                has_image = True
            
            # キーワードマッチング
            if expected_str and expected_str in cell_str:
                return "OK", has_image
    
    return "NG", has_image


def main():
    """メイン処理"""
    
    print("試験項目書とエビデンスの整合性チェックを開始します...")
    print(f"実施日: {datetime.now().strftime('%Y-%m-%d')}")
    print()
    
    # ファイルの存在確認
    if not os.path.exists(TEST_ITEMS_FILE):
        print(f"エラー: {TEST_ITEMS_FILE} が見つかりません")
        return False
    
    if not os.path.exists(EVIDENCE_FILE):
        print(f"エラー: {EVIDENCE_FILE} が見つかりません")
        return False
    
    try:
        # エビデンスを読み込む
        print(f"エビデンスを読み込み中: {EVIDENCE_FILE}")
        evidence_wb = openpyxl.load_workbook(EVIDENCE_FILE)
        
        if EVIDENCE_SHEET not in evidence_wb.sheetnames:
            print(f"エラー: シート '{EVIDENCE_SHEET}' が見つかりません")
            print(f"利用可能なシート: {evidence_wb.sheetnames}")
            return False
        
        evidence_ws = evidence_wb[EVIDENCE_SHEET]
        evidence_data = []
        
        # エビデンスのすべてのデータを読み込む
        for row in evidence_ws.iter_rows(values_only=True):
            evidence_data.append(row)
        
        print(f"エビデンスデータを読み込みました ({len(evidence_data)} 行)")
        print()
        
        # 試験項目書を読み込む
        print(f"試験項目書を読み込み中: {TEST_ITEMS_FILE}")
        test_wb = openpyxl.load_workbook(TEST_ITEMS_FILE)
        
        if TEST_ITEMS_SHEET not in test_wb.sheetnames:
            print(f"エラー: シート '{TEST_ITEMS_SHEET}' が見つかりません")
            print(f"利用可能なシート: {test_wb.sheetnames}")
            return False
        
        test_ws = test_wb[TEST_ITEMS_SHEET]
        
        # カラムを特定
        # 「期待値」→「確認内容」、「実施状況」→「結果」「実施日」「実施者」を見つける
        header_row = None
        expected_col = None
        result_col = None
        date_col = None
        implementer_col = None
        
        # ヘッダー行を探す
        for row_idx in range(1, test_ws.max_row + 1):
            headers = []
            for col in range(1, test_ws.max_column + 1):
                headers.append(test_ws.cell(row=row_idx, column=col).value)
            
            # 「確認内容」を含む列を探す
            if any('確認内容' in str(h or '') for h in headers):
                header_row = row_idx
                
                for col_idx, header in enumerate(headers, 1):
                    if header and '確認内容' in str(header):
                        expected_col = col_idx
                    elif header and '結果' in str(header) and '実施状況' in str(test_ws.cell(row=row_idx, column=col_idx).value or ''):
                        result_col = col_idx
                    elif header and '実施日' in str(header):
                        date_col = col_idx
                    elif header and '実施者' in str(header):
                        implementer_col = col_idx
                
                break
        
        if not header_row:
            print("ヘッダー行が見つかりません")
            return False
        
        print(f"ヘッダー行: {header_row}")
        print(f"確認内容列: {expected_col}, 結果列: {result_col}, 実施日列: {date_col}, 実施者列: {implementer_col}")
        print()
        
        # データ行を処理
        current_date = datetime.now().strftime('%Y-%m-%d')
        updated_count = 0
        
        for row_idx in range(header_row + 1, test_ws.max_row + 1):
            expected_content = test_ws.cell(row=row_idx, column=expected_col).value if expected_col else None
            
            if not expected_content:
                continue
            
            # エビデンスをチェック
            result, has_image = check_evidence_coverage(expected_content, evidence_data)
            
            # 結果を記載
            if result_col:
                result_text = f"※目視で確認してください\n{result}" if has_image else result
                test_ws.cell(row=row_idx, column=result_col).value = result_text
            
            # 実施日を記載
            if date_col:
                test_ws.cell(row=row_idx, column=date_col).value = current_date
            
            # 実施者を記載
            if implementer_col:
                test_ws.cell(row=row_idx, column=implementer_col).value = IMPLEMENTER_NAME
            
            updated_count += 1
            status = "OK" if result == "OK" else "NG"
            image_note = " (画像あり)" if has_image else ""
            print(f"行 {row_idx}: {result}{image_note}")
        
        # 結果を保存
        print()
        print(f"{updated_count} 行を更新しました")
        test_wb.save(TEST_ITEMS_FILE)
        print(f"試験項目書を保存しました: {TEST_ITEMS_FILE}")
        print()
        print("チェック完了！")
        
        return True
    
    except Exception as e:
        print(f"エラーが発生しました: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
